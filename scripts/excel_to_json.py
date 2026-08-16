#!/usr/bin/env python3
"""
Prevodník: cviky-sablona.xlsx  ->  data/temy/tema_90_excel.json  -> exercises.json

    python3 scripts/excel_to_json.py [cesta_k_xlsx]
    python3 scripts/build_db.py && python3 scripts/inject.py

Excel je zdroj pravdy pre cviky nahrané cez šablónu. Ostatné sady (elitná
a tematické) zostávajú nedotknuté — build_db.py ich všetky zloží dokopy
a skontroluje, či má každý cvik vyplnené povinné polia.
"""
import sys, json, os

COLS = ["ID","Názov","Časť","Zručnosť","Ďalšie zručnosti","Téma","Vek od","Vek do",
        "Hráči","Priestor","Intenzita","Minúty","Pomôcky","Prečo","Rozostavenie",
        "Priebeh","Podmienky","Sťaženie","Zjednodušenie","Na čo dbať","Dávkovanie","Úroveň"]

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "cviky-sablona.xlsx"
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "Cviky" not in wb.sheetnames:
        print("Chyba: hárok 'Cviky' nenájdený."); sys.exit(1)
    ws = wb["Cviky"]
    hdr = {str(ws.cell(row=1, column=c).value).strip(): c
           for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}
    def cell(r, name):
        c = hdr.get(name)
        if not c: return ""
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    out, missing = [], []
    n = 0
    for r in range(2, ws.max_row + 1):
        name = cell(r, "Názov")
        if not name: continue
        n += 1
        od, do = cell(r,"Vek od"), cell(r,"Vek do")
        mins = cell(r,"Minúty")
        extra = [s.strip() for s in cell(r,"Ďalšie zručnosti").replace(";", ",").split(",") if s.strip()]
        ex = {
            "id": cell(r,"ID") or f"X{n:03d}",
            "name": name,
            "theme": cell(r,"Téma"),
            "phase": cell(r,"Časť"),
            "age": f"{od}–{do}" if od and do else (od or do),
            "players": cell(r,"Hráči"),
            "space": cell(r,"Priestor"),
            "time": (mins + "´") if mins else "",
            "gear": cell(r,"Pomôcky"),
            "why": cell(r,"Prečo"),
            "setup": cell(r,"Rozostavenie"),
            "steps": cell(r,"Priebeh"),
            "constraints": cell(r,"Podmienky"),
            "progression": cell(r,"Sťaženie"),
            "regression": cell(r,"Zjednodušenie"),
            "coach": cell(r,"Na čo dbať"),
            "load": cell(r,"Dávkovanie"),
            "level": cell(r,"Úroveň") or "foundation",
            "intensity": cell(r,"Intenzita"),
        }
        if cell(r,"Zručnosť"):
            ex["skills"] = [cell(r,"Zručnosť")] + extra
        chyba = [k for k in ("theme","phase","why","setup","steps","constraints",
                             "progression","regression","coach","load") if not ex[k]]
        if chyba: missing.append(f"  riadok {r} „{name}“ — nevyplnené: {', '.join(chyba)}")
        out.append(ex)

    os.makedirs(os.path.join(root,"data","temy"), exist_ok=True)
    dest = os.path.join(root,"data","temy","tema_90_excel.json")
    with open(dest, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(out)} cvikov -> data/temy/tema_90_excel.json")
    if missing:
        print(f"\n⚠ {len(missing)} cvikov nemá vyplnené všetky polia — build_db.py ich odmietne:")
        for m in missing[:20]: print(m)
        print("\nDoplň ich v Exceli a spusti skript znova.")
    else:
        print("Ďalej: python3 scripts/build_db.py && python3 scripts/inject.py")

if __name__ == "__main__":
    main()
